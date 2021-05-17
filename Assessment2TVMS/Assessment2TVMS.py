# -*- coding: utf-8 -*-
import openpyxl,sys,argparse,re
from pefile import fast_load
TVMS_HEADERS = {"A":"弱點發現時間",\
                "B":"弱點所在網路位址",\
                "C":"弱點所在網路埠號",\
                "D":"檔案名稱/URL",\
                "E":"弱點所在之網路協定",\
                "F":"弱點名稱",\
                "G":"弱點嚴重性或合規檢測結果",\
                "H":"弱點類別",\
                "I":"弱點CVE ID清單",\
                "J":"評估工具原廠之弱點編號",\
                "K":"弱點說明",\
                "L":"弱點意見",\
                "M":"弱點修補建議",\
                "N":"弱點證據描述",\
                "O":"類型(弱點或合規)"}
VUL_NAME = {"D":"作業系統安全性更新編號",\
            "E":"Office應用程式安全性更編號",\
            "F":"Adobe Reader更新", \
            "G":"Adobe Flash Player更新", \
            "H":"Java更新",\
            "I":"防毒軟體",\
            "J":"惡意程式檢測結果"}
UNUPDATED = "未更新"
UPDATED = "已更新至最新"
KBNOTFOUND = "kbid is not found"
MALWARENOTFOUND = "未發現惡意程式"
FAILED = "不符合"
PASSED = "符合"
TOOL_NAME = "神網資產管理系統健診說明"
CHECK_TYPE = "合規"
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("-r", "--read", help = "來源Excel檔案路徑")
    parser.add_argument("-o", "--output", help = "輸出TVMS檔案路徑")
    args = parser.parse_args()
    INPUT_FILE = args.read
    OUTPUT_FILE = args.output
    try:
        input_file = openpyxl.load_workbook(INPUT_FILE,read_only=False,data_only=True)
    except :
        print(f'Unable to open {INPUT_FILE}--', sys.exc_info()[0])
        raise
    if len(input_file.worksheets)==0:
        sys.exit(f"檔案 {INPUT_FILE} 內沒有工作表")
    try:
        output_file = open(OUTPUT_FILE,encoding='utf-8-sig',mode='w')
    except :
        print(f'Unable to open {OUTPUT_FILE}--', sys.exc_info()[0])
        raise
    output_file.write(",".join(list(TVMS_HEADERS.values()))+"\n")
    ws = input_file.active
    for cell in ws['D']:
        if cell.row == 1:
            continue # skip the first row
        else:
            if (UNUPDATED in cell.value) or (KBNOTFOUND in cell.value):
                result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['D']
                result_list['G'] = FAILED
                
                m = re.search("(?<=:).+(?=;)", cell.value)
                if m: # UNUPDATED
                    s = ";".join(["KB"+i.strip() for i in m.group(0).split(";")])
                    result_list['K'] = "未更新" + s
                
                m = re.search(KBNOTFOUND, cell.value)
                if m: # KBNOTFOUND
                    result_list['K'] = m.group(0)
                
                result_list['J'] = TOOL_NAME
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output_file.write(",".join(list(result_list.values()))+"\n")
    
    for cell in ws['E']:
        if cell.row == 1:
            continue # skip the first row
        else:
            if UNUPDATED in cell.value:
                result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['E']
                result_list['G'] = FAILED
                
                m = re.search("(?<=:).+(?=;)", cell.value)
                if m: # UNUPDATED
                    s = ";".join(["KB"+i.strip() for i in m.group(0).split(";")])
                    result_list['K'] = "未更新" + s
                
                m = re.search(KBNOTFOUND, cell.value)
                if m: # KBNOTFOUND
                    result_list['K'] = m.group(0)
                
                result_list['J'] = TOOL_NAME
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output_file.write(",".join(list(result_list.values()))+"\n")
    
    for cell in ws['F']:
        if cell.row == 1:
            continue # skip the first row
        else:
            if UNUPDATED in cell.value:
                result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['F']
                result_list['G'] = FAILED
                result_list['K'] = UNUPDATED
                result_list['J'] = TOOL_NAME
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output_file.write(",".join(list(result_list.values()))+"\n")
    
    for cell in ws['G']:
        if cell.row == 1:
            continue # skip the first row
        else:
            if UNUPDATED in cell.value:
                result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['G']
                result_list['G'] = FAILED
                result_list['K'] = UNUPDATED
                result_list['J'] = TOOL_NAME
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output_file.write(",".join(list(result_list.values()))+"\n")
    
    for cell in ws['H']:
        if cell.row == 1:
            continue # skip the first row
        else:
            if UNUPDATED in cell.value:
                result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['H']
                result_list['G'] = FAILED
                result_list['K'] = UNUPDATED
                result_list['J'] = TOOL_NAME
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output_file.write(",".join(list(result_list.values()))+"\n")
    
    for cell in ws['I']:
        if cell.row == 1:
            continue # skip the first row
        else:
            if not (UPDATED in cell.value):
                result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['I']
                result_list['G'] = FAILED
                result_list['K'] = cell.value
                result_list['J'] = TOOL_NAME
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output_file.write(",".join(list(result_list.values()))+"\n")
    
    for cell in ws['J']:
        if cell.row == 1:
            continue # skip the first row
        else:
            if MALWARENOTFOUND not in cell.value:
                result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['J']
                result_list['G'] = FAILED
                result_list['K'] = UNUPDATED
                result_list['J'] = TOOL_NAME
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output_file.write(",".join(list(result_list.values()))+"\n")
    
    
    input_file.close()
    output_file.close()
if __name__ == '__main__':
    try:
        main()
    except:
        pass
