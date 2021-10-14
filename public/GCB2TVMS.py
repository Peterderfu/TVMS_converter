# -*- coding: utf-8 -*-
import openpyxl,sys,argparse,os,glob
from email.policy import default
ARG_FILE = "arg.txt"
TVMS_TEMP = r"C:\Users\1911010\git\TVMS_converter\GCB2TVMS\doc\Windows弱點類型對照表.xlsx"
TVMS_HEADERS = {"A":"弱點發現時間",\
                "B":"弱點所在網路位址",\
                "C":"弱點所在網路埠號",\
                "D":"檔案名稱/URL",\
                "E":"弱點所在之網路協定",\
                "F":"弱點名稱",\
                "G":"弱點嚴重性或合規檢測結果",\
                "H":"弱點類別",\
                "I":"弱點CVE ID清單",\
                "J":"評估工具原廠之弱點編號",\
                "K":"弱點說明",\
                "L":"弱點意見",\
                "M":"弱點修補建議",\
                "N":"弱點證據描述",\
                "O":"類型(弱點或合規)"}
FAILED = "不符合"
PASSED = "符合"
TOOL_NAME = "神網資產管理系統健診說明"
CHECK_TYPE = "合規"
GCB_AUDIT = 'GCB合規檢測'
DESCRIPTION = "詳細報告請參考附件"
SUGGESTION = "可依照各機關實際導入狀況 自行評估導入範圍，如無法全數導入請填寫原因或納入例外管理"

MAP2TVMS = {"B":"位址",\
            "G":"結果",\
            "H":"組態類型",\
            "N":"主機設定值"}
def getRecSeq(count):
    global RECORD_PREFIX
    return "{}{:0>10d}".format(RECORD_PREFIX,count)
def main():
    global RECORD_PREFIX
    parser = argparse.ArgumentParser()
    parser.add_argument("-r", "--read", help = "來源Excel檔案路徑")
    parser.add_argument("-s", "--sheet", default = "主機分類總覽", help = "Excel工作表名稱")
    parser.add_argument("-o", "--output", help = "輸出TVMS檔案路徑")
#     parser.add_argument("-n", "--number",help = "評估工具原廠之弱點編號",default="北市府110年資安健診-軟體更新-M")
    parser.add_argument("-n", "--number",help = "評估工具原廠之弱點編號",default="")
    parser.add_argument("-p", "--simple",help = "簡略弱點說明與修補建議",action="store_true")
    parser.add_argument("-t", "--template",help = "TVMS template",default=TVMS_TEMP)
    parser.add_argument("-x","--invalid",help = "僅列出不符合項目",action="store_true",default=False)
    
    args = parser.parse_args()
    INPUT_FILE          = args.read
    OUTPUT_DIR          = args.output    
    SHEET               = args.sheet
    RECORD_PREFIX       = args.number
    SIMPLE              = args.simple
    TEMPLATE            = args.template
    ONLYINVALID         = args.invalid
    POWERSHELL_LOG_DIR  = os.path.join(os.path.dirname(INPUT_FILE),"winserver")
    OUTPUT_FILE         = os.path.join(OUTPUT_DIR,"TVMS_Winserv_" + os.path.splitext(os.path.basename(INPUT_FILE))[0]+".csv")
    
    
    try:  #open input Excel file
        input_file = openpyxl.load_workbook(INPUT_FILE,read_only=False,data_only=True)        
    except :
        print(f'Unable to open {INPUT_FILE}--', sys.exc_info()[0])
        raise
    
    if not input_file[SHEET]:
        sys.exit(f"檔案 :{INPUT_FILE} 內沒有工作表：{SHEET}，請以-s設定")
    ws = input_file[SHEET]  # open desired worksheet in Excel
    
    try: #open TVMS template file
        template   = openpyxl.load_workbook(TEMPLATE,read_only=False,data_only=True)
    except:
        print(f'Unable to open {TEMPLATE}--', sys.exc_info()[0])
        raise
    
    template_ws = dict(zip(["2008","2012","2016","2019","Windows 10"],[template["2008"],template["2012"],template["2016"],template["2019"],template["Windows 10"]]))

    try: # open the output file for writing
        output_file = open(OUTPUT_FILE,encoding='utf-8-sig',mode='w')
    except :
        print(f'Unable to open {OUTPUT_FILE}--', sys.exc_info()[0])
        raise
    output_file.write(",".join(list(TVMS_HEADERS.values()))+"\n") # the 1st line is the fields of TVMS  
    
    input_header = ws[1] #retrieve the 1st row of input row
    map2TVMS = {}
    IP_HOST = {}
    for cell in input_header:
        try:
            map2TVMS[list(MAP2TVMS.keys())[list(MAP2TVMS.values()).index(cell.value)]] = cell.column_letter
        except ValueError: #cell.value is not a value in MAP2TVMS
            continue
    for cell in ws['R']: # lock to the column of result
        if cell.row == 1:
            continue # skip the first row
        else:
            IP_HOST.update({ws[f'B{cell.row}'].value:ws[f'A{cell.row}'].value})
            
            if ONLYINVALID and (cell.value == PASSED): # only pick up the rows with failed result
                continue
            else:
                result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                for k,v in map2TVMS.items(): # map input file columns to TVMS column
                    try:
                        result_list[k] = ws[f'{v}{cell.row}'].value.replace("\n","").strip(" \n").replace(",","_")
                    except:
                        pass
                
                result_list['F'] = GCB_AUDIT
                if not (cell.value == PASSED):
                    result_list['G'] = FAILED
                else:
                    result_list['G'] = PASSED
                result_list['O'] = CHECK_TYPE
                if SIMPLE:
                    result_list['K'] = DESCRIPTION
                    result_list['M'] = SUGGESTION
                else:
                    result_list['K'] = ws[f'M{cell.row}'].value.replace("\n","").strip(" \n").replace(",","_")
                    result_list['M'] = ws[f'P{cell.row}'].value.replace("\n","").strip(" \n").replace(",","_")
                if len(RECORD_PREFIX)==0:
                    result_list['J'] ="/".join([ws[f'J{cell.row}'].value,ws[f'K{cell.row}'].value])
                else:
                    result_list['J'] = getRecSeq(cell.row-1)
                
                for k in template_ws.keys():
                    if k in ws[f'C{cell.row}'].value:
                        OS = k
                        break
                
                configName = ws[f'M{cell.row}'].value
                
                for c in template_ws[OS]['D']:
                    if configName in c.value:
                        result_list['H'] =  template_ws[OS][f'C{c.row}'].value
                        break
                
                result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                output_file.write(",".join(list(result_list.values()))+"\n")
    
    
    powershell_log = glob.glob(POWERSHELL_LOG_DIR+"/*.log")
    for l in powershell_log:
        with open(l,mode='r',encoding='utf8') as ff:
            lines = ff.readlines()
            for line in lines:
                if ONLYINVALID and line.startswith('+'):
                    continue
                result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                if "資料執行防止" in line:
                    result_list['H'] = "Data Execution Prevention"
                    result_list['K'] = "為所有的Windows程式和服務開啟資料執行防止(DEP)"
                    result_list['M'] = "為所有的Windows程式與服務開啟DEP"
                    if line.startswith('-'):
                        result_list['N'] = "未開啟DEP"
                elif ("139" in line):
                    result_list['H'] = "設定輸入輸出規則(關閉139/445/5985)"
                    result_list['K'] = "關閉網路芳鄰與遠端管理服務WinRM 服務"
                    result_list['M'] = "關閉139_445_5985"
                    if line.startswith('-'):
                        result_list['N'] = "未關閉139，且未限制IP"
                elif ("445" in line):
                    result_list['H'] = "設定輸入輸出規則(關閉139/445/5985)"
                    result_list['K'] = "關閉網路芳鄰與遠端管理服務WinRM 服務"
                    result_list['M'] = "關閉139_445_5985"
                    if line.startswith('-'):
                        result_list['N'] = "未關閉445，且未限制IP"
                elif ("5985" in line):
                    result_list['H'] = "設定輸入輸出規則(關閉139/445/5985)"
                    result_list['K'] = "關閉網路芳鄰與遠端管理服務WinRM 服務"
                    result_list['M'] = "關閉139_445_5985"
                    if line.startswith('-'):
                        result_list['N'] = "未關閉5985，且未限制IP"                    
                elif "ICMP" in line:
                    result_list['H'] = "Ping(ICMP)回應開啟"
                    result_list['K'] = "開啟PING回應"
                    result_list['M'] = "開啟PING回應"
                    if line.startswith('-'):
                        result_list['N'] = "未開啟PING回應"
                elif "遠端桌面" in line:
                    result_list['H'] = "遠端桌面限制IP連線"
                    result_list['K'] = "限制遠端桌面來源IP連線"
                    result_list['M'] = "限VPN連線或限縮來源IP"
                    if line.startswith('-'):
                        result_list['N'] = "未限制來源IP"
                elif "螢幕保護" in line:
                    result_list['H'] = "螢幕保護裝置時間設定"
                    result_list['K'] = "螢幕保護裝置等待時間"
                    result_list['M'] = "螢幕保護裝置等待時間應小於15分鐘"
                    if line.startswith('-'):
                        result_list['N'] = "螢幕保護裝置等待時間大於15分鐘或未設定"
                elif "Guest" in line:
                    result_list['H'] = "Guest帳號關閉"
                    result_list['K'] = "Guest帳號關閉"
                    result_list['M'] = "Guest帳號應關閉"
                    if line.startswith('-'):
                        result_list['N'] = "Guest帳號未關閉"
                if line.startswith('+'):
                    result_list['N'] = "設定正確"
                    result_list['G'] = PASSED
                else:
                    result_list['G'] = FAILED
                    
                result_list['B'] = IP_HOST[os.path.basename(l).rsplit('_',maxsplit=1)[0]].replace(",","_")
                result_list['F'] = GCB_AUDIT
                
                result_list['O'] = CHECK_TYPE 
                result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                output_file.write(",".join(list(result_list.values()))+"\n")
    
    input_file.close()
    output_file.close()
if __name__ == '__main__':
    try:
        main()
    except:
        print(f'Unable to run this program--', sys.exc_info()[0])
        raise
