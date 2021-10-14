# -*- coding: utf-8 -*-
import openpyxl,sys,argparse,re,os,csv
from os.path import join
MERGED_HEADER = {'A':'Hostname',\
                 'B':'IP',\
                 'C':'OS version',\
                 'D':'Office',\
                 'E':'Adobe Reader',\
                 'F':'Adobe Reader latest',\
                 'G':'Adobe Reader status',\
                 'H':'Flash Player',\
                 'I':'Flash Player latest',\
                 'J':'Flash Player status',\
                 'K':'Java',\
                 'L':'Java latest',\
                 'M':'Java status',\
                 'N':'Antivirus'}
TVMS_HEADERS = {"A":"弱點發現時間",\
                "B":"弱點所在網路位址",\
                "C":"弱點所在網路埠號",\
                "D":"檔案名稱/URL",\
                "E":"弱點所在之網路協定",\
                "F":"弱點名稱",\
                "G":"弱點嚴重性或合規檢測結果",\
                "H":"弱點類別",\
                "I":"弱點CVE ID清單",\
                "J":"評估工具原廠之弱點編號",\
                "K":"弱點說明",\
                "L":"弱點意見",\
                "M":"弱點修補建議",\
                "N":"弱點證據描述",\
                "O":"類型(弱點或合規)"}
VUL_NAME = {"D":"作業系統安全性更新編號",\
            "E":"Office應用程式安全性更編號",\
            "F":"Adobe Reader未更新", \
            "G":"Flash Player未更新", \
            "H":"Java未更新",\
            "I":"防毒軟體未更新",\
            "J":"防毒軟體未安裝",\
            "K":"防毒軟體病毒碼未更新",\
            "L":"Office軟體版本過舊",\
            "M":"Windows作業系統版本過舊"}
UNUPDATED = "未更新"
UPDATED = "已更新至最新"
KBNOTFOUND = "kbid is not found"
MALWARENOTFOUND = "未發現惡意程式"
FAILED = "不符合"
PASSED = "符合"
CHECK_TYPE = "合規"
MBSA = {"ADOBE":"adobe_big5.csv",\
        "FLASHPLAYER":"flashplayer_big5.csv",\
        "JAVA":"java_big5.csv",\
        "ANTIVIRUS":"antivirus_big5.csv"}
R0=R1=R2=R3=R4=R5=R6=None
AV_BASE_VER = ''
RECORD_PREFIX = ''
KEY_APEX = 'Apex'
KEY_OFFICESCAN = 'OfficeScan'
count = 0
def is_office_phaseout(s):
    pattern = 'office.+(95|97|2003|2007|2010)'
    return re.search(pattern,s,re.I)

def is_windows_phaseout(s):
    pattern = 'Windows.+(7|Vista|XP|ME|2000|98|NT|95|2008|2003)'
    return re.search(pattern, s)

def getRecSeq(count):
    global RECORD_PREFIX
    return "{}{:0>10d}".format(RECORD_PREFIX,count)
def csv2exel(csvfile):
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(csvfile,newline='') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
    return wb
def init_input(input_dir):
    global R0,R1,R2,R3,R4,R5,R6
     #檢驗資料夾結構     
    if not os.path.isdir(input_dir): 
        sys.exit(f"資料夾:{input_dir} 不存在")
    subdirs = sorted(os.listdir(input_dir))
    if len(subdirs) == 0:
        sys.exit(f"{input_dir}內無資料夾")
    if not (subdirs[0].startswith("01_") and \
            subdirs[1].startswith("02_") and \
            subdirs[2].startswith("03_") and \
            subdirs[3].startswith("04_")):
        sys.exit(f"{input_dir} 內資料夾名稱不符-請用 以下名稱\n \
                    01_伺服器更新報表\n \
                    02_MBSA版本報表\n \
                    03_軟體版本更新報表\n \
                    04_作業系統版本報表")
    subdirs = [os.path.normpath(join(input_dir,d)) for d in subdirs]
    
    #開啟伺服器更新報表
    R0 = os.listdir(subdirs[0])
    try:
        if (len(R0) == 0):
            sys.exit(f"\"{subdirs[0]}\" 資料夾內沒有更新報表檔案")
        if not (len(R0) == 1):
            sys.exit(f"\"{subdirs[0]}\" 資料夾內應僅有1個更新報表檔案")
    except SystemExit as e:
        print(str(e))
        exit
    R0 = os.path.normpath(join(subdirs[0],R0[0]))
    [_,ext]=os.path.splitext(R0)
    if ext=='.csv':
        R0 = csv2exel(R0)
    else:
        try:
            R0 = openpyxl.load_workbook(R0,read_only=False,data_only=True)
        except :
            print(f'無法開啟伺服器更新報表檔案 {R0}--', sys.exc_info()[0])
            exit
        
    #開啟MBSA版本報表
    temp = os.listdir(subdirs[1])
    try:
        if (len(temp) == 0):
            sys.exit(f"\"{subdirs[1]}\" 資料夾內沒有MBSA版本報表檔案")
        for t in list(MBSA.values()):
            if not (t in temp):
                sys.exit(f"\"{subdirs[1]}\" 資料夾內沒有MBSA版本報表檔案: {t}")
    except SystemExit as e:
        print(str(e))
        exit
    try:
        R1 = os.path.normpath(join(subdirs[1],MBSA['ADOBE']))
        [_,ext]=os.path.splitext(R1)
        if ext=='.csv':
            R1 = csv2exel(R1)
        else:            
            R1 = open(R1,encoding='big5',newline='')
        R2=os.path.normpath(join(subdirs[1],MBSA['FLASHPLAYER']))
        [_,ext]=os.path.splitext(R2)
        if ext=='.csv':
            R2 = csv2exel(R2)
        else:    
            R2 = open(R2,encoding='big5',newline='')
        R3 = os.path.normpath(join(subdirs[1],MBSA['JAVA']))
        [_,ext]=os.path.splitext(R3)
        if ext=='.csv':
            R3 = csv2exel(R3)
        else:    
            R3 = open(R3,encoding='big5',newline='')
        R4 = os.path.normpath(join(subdirs[1],MBSA['ANTIVIRUS']))
        [_,ext]=os.path.splitext(R4)
        if ext=='.csv':
            R4 = csv2exel(R4)
        else:    
            R4 = open(R4,encoding='big5',newline='')
    except :
        print(f'無法開啟MBSA版本報表檔案 --', sys.exc_info()[0])
        exit
    
    #開啟軟體版本更新報表
    R5 = os.listdir(subdirs[2])
    try:
        if (len(R5) == 0):
            sys.exit(f"\"{subdirs[5]}\" 資料夾內沒有軟體版本報表檔案")
    except SystemExit as e:
        print(str(e))
        exit
    R5 = [os.path.normpath(join(subdirs[2],f)) for f in R5]
    
    #開啟作業系統版本報表
    R6 = os.listdir(subdirs[3])
    try:
        if (len(R6) == 0):
            sys.exit(f"\"{subdirs[3]}\" 資料夾內沒有作業系統版本檔案")
        if not (len(R6) == 1):
            sys.exit(f"\"{subdirs[3]}\" 資料夾內應僅有1個作業系統版本檔案")
    except SystemExit as e:
        print(str(e))
        exit    
    R6 = os.path.normpath(join(subdirs[3],R6[0]))
    try:
        R6 = openpyxl.load_workbook(R6,read_only=False,data_only=True)
    except :
        print(f'無法開啟作業系統版本報表檔案 {R6}--', sys.exc_info()[0])
        exit
def process_R0(R0):
    global count
    output = []
    ws = R0.active
    for cols in ws['D:E']: #依序讀取D,E欄的儲存格
        for cell in cols:
            if cell.row == 1:
                continue # skip the first row
            else:
                if (UNUPDATED in cell.value) or (KBNOTFOUND in cell.value): #儲存格中有  【未更新】 或 【kbid is not found】
                    result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                    IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                    result_list['F'] = VUL_NAME[cell.column_letter]
                    result_list['G'] = FAILED
                    
                    m = re.search("(?<=:).+(?=;)", cell.value)
                    if m: # UNUPDATED
                        s = ";".join(["KB"+i.strip() for i in m.group(0).split(";")])
                        result_list['K'] = "目前版本：未更新" + s
                        result_list['M'] = "更新至" + s + "或以上版本"
                    
                    m = re.search(KBNOTFOUND, cell.value)
                    if m: # KBNOTFOUND
                        result_list['K'] = m.group(0)
                        result_list['M'] = "Windows update安全性更新至最新"
                    
                    result_list['O'] = CHECK_TYPE
                    for ip in IP:
                        result_list['J'] = getRecSeq(count+1)
                        result_list['B'] = ip
                        result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                        output.append(",".join(list(result_list.values())))
                        count = count + 1
                    result_list = None
    return output

def process_R1(R1):
    global count
    output = []
    ws = R1.active
    for cell in ws['F']:
        if cell.row == 1:
            continue  # skip the first row
        else:
            if UNUPDATED in cell.value:
                result_list = dict(zip(TVMS_HEADERS.keys(), [''] * len(TVMS_HEADERS.keys())))
                IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['F']
                result_list['G'] = FAILED
                result_list['K'] = "目前版本：Adobe Reader " + ws[f'E{cell.row}'].value
                result_list['M'] = "更新至{}或以上版本".format(ws[f'D{cell.row}'].value)
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['J'] = getRecSeq(count+1)
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output.append(",".join(list(result_list.values())))
                    count = count + 1
                result_list = None
    return output
def process_R2(R2):
    global count
    output = []
    ws = R2.active
    for cell in ws['F']:
        if cell.row == 1:
            continue  # skip the first row
        else:
            if UNUPDATED in cell.value:
                result_list = dict(zip(TVMS_HEADERS.keys(), [''] * len(TVMS_HEADERS.keys())))
                IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['G']
                result_list['G'] = FAILED
                result_list['K'] = "目前版本：Flash Player " + ws[f'E{cell.row}'].value
                result_list['M'] = "建議移除軟體"
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['J'] = getRecSeq(count+1)
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output.append(",".join(list(result_list.values())))
                    count = count + 1
                result_list = None
    return output
def process_R3(R3):
    global count
    output = []
    ws = R3.active
    for cell in ws['F']:
        if cell.row == 1:
            continue  # skip the first row
        else:
            if UNUPDATED in cell.value:
                result_list = dict(zip(TVMS_HEADERS.keys(), [''] * len(TVMS_HEADERS.keys())))
                IP = ws[f'C{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['H']
                result_list['G'] = FAILED
                result_list['K'] = "目前版本：Java " + ws[f'E{cell.row}'].value
                result_list['M'] = "更新至{}或以上版本".format(ws[f'D{cell.row}'].value)
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['J'] = getRecSeq(count+1)
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output.append(",".join(list(result_list.values())))
                    count = count + 1
                result_list = None
    return output
def process_R4(R4):
    global count
    output = []
    cols = ['uuid','hostname','ip','latest','avname','current','status']
    colMap = dict(zip(cols,[None]*len(cols)))
    ws = R4.active
    for i in ws['1']:
        colMap[i.value]=i.column_letter
    for cell in ws[f'{colMap["status"]}']: #check "防毒軟體病毒碼未更新"
        if cell.row == 1:
            continue  # skip the first row
        else:
            if UNUPDATED in cell.value:
                result_list = dict(zip(TVMS_HEADERS.keys(), [''] * len(TVMS_HEADERS.keys())))
                IP = ws[f'{colMap["ip"]}{cell.row}'].value.strip(";").split(";")
                result_list['F'] = VUL_NAME['K']
                result_list['G'] = FAILED
                result_list['K'] = "目前版本： " + ws[f'{colMap["current"]}{cell.row}'].value
                result_list['M'] = "更新至{}或以上版本".format(ws[f'{colMap["latest"]}{cell.row}'].value)
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['J'] = getRecSeq(count+1)
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output.append(",".join(list(result_list.values())))
                    count = count + 1
                result_list = None
    return output
def process_R5(R4,R5):
    global count
    output = []
    cols = ['uuid','hostname','ip','latest','avname','current','status']
    colMap = dict(zip(cols,[None]*len(cols)))
    antivirus = R4.active
    for i in antivirus['1']:
        colMap[i.value]=i.column_letter
    IP_STATUS = dict(zip([c.value for c in antivirus[f'{colMap["ip"]}']],[c.value for c in antivirus[f'{colMap["status"]}']]))
    for f in R5:
        m = re.search('(?<=_)\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}.*(?=_)', f)
        if m:  # filename pattern "_[IPv4]_" found 
            IP = m.group(0).split(',')
        else:
            sys.exit('No IP pattern found in {f}')
            raise

        try:
            r = openpyxl.load_workbook(f,read_only=False,data_only=True)
        except :
            print(f'無法開啟軟體版本更新報表檔案 --', sys.exc_info()[0])
            exit
        ws = r.active
        av_ver = {} #AV versions
        office_ver = [] #Office versions
        for cell in ws['B']: #check "防毒軟體未安裝"  & "防毒軟體未更新" & "Office軟體版本過舊"
            if cell.row == 1:
                continue  # skip the first row
            else:
                if ("Trend Micro Apex One Security Agent".lower() in cell.value.lower()):
                    av_ver.update({KEY_APEX:ws[f'C{cell.row}'].value})
                if ("OfficeScan" in cell.value.lower()):
                    av_ver.update({KEY_OFFICESCAN:ws[f'C{cell.row}'].value})
                if ("office" in cell.value.lower()) and not ("OfficeScan" in cell.value.lower()):
                    office_ver.append(cell.value)
                    
        if len(av_ver) == 0: # check "防毒軟體未安裝"
            result_list = dict(zip(TVMS_HEADERS.keys(), [''] * len(TVMS_HEADERS.keys())))
            result_list['F'] = VUL_NAME['J']
            result_list['G'] = FAILED
            result_list['K'] = "目前版本：未發現防毒軟體"
            result_list['M'] = f'安裝本府趨勢防毒軟體{AV_BASE_VER}版本'
            result_list['O'] = CHECK_TYPE
        else:
            result_list = dict(zip(TVMS_HEADERS.keys(), [''] * len(TVMS_HEADERS.keys())))
            result_list['F'] = VUL_NAME['I']
            result_list['O'] = CHECK_TYPE
            curver = ''
            curver = av_ver.get(KEY_APEX)
            if not curver:
                curver = av_ver.get(KEY_OFFICESCAN)
             
            if AV_BASE_VER > curver: # get Apex'ver insted of Officescan's ver and compare version smaller than required
                result_list['G'] = FAILED
                result_list['K'] = "目前版本： " + curver
                result_list['M'] = f'安裝本府趨勢防毒軟體{AV_BASE_VER}版本'
            else:
                result_list['G'] = PASSED
                result_list['K'] = "已更新至最新版本"
                result_list['M'] = "已更新至最新版本，無修補建議"
                
        for i in list(IP_STATUS.keys()):
            if result_list['K'] == "目前版本：未發現防毒軟體":
                if ((UPDATED in IP_STATUS[i]) or (UNUPDATED in IP_STATUS[i])):
                    result_list = None
                    break
        if result_list:
            for ip in IP:
                result_list['J'] = getRecSeq(count+1)
                result_list['B'] = ip
                result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                output.append(",".join(list(result_list.values())))
                count = count + 1
            result_list = None
        for v in office_ver:
            result_list = dict(zip(TVMS_HEADERS.keys(), [''] * len(TVMS_HEADERS.keys())))
            result_list['F'] = VUL_NAME['L']
            result_list['O'] = CHECK_TYPE
            if is_office_phaseout(v):
                result_list['G'] = FAILED
                result_list['K'] = "目前版本： " + v
                result_list['M'] = "更新至Microsoft Office 2013或以上"
            else:
                result_list['G'] = PASSED
                result_list['K'] = "已更新至最新版本"
                result_list['M'] = "已更新至最新版本，無修補建議"
        if result_list:
            for ip in IP:
                result_list['J'] = getRecSeq(count+1)
                result_list['B'] = ip
                result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                output.append(",".join(list(result_list.values())))
                count = count + 1
            result_list = None
    return output
def process_R6(R6):
    global count
    output = []
    ws = R6.active
    for cell in ws['E']: #check "Windows作業系統版本過舊"
        if cell.row == 1:
            continue  # skip the first row
        else:
            if is_windows_phaseout(cell.value):
                result_list = dict(zip(TVMS_HEADERS.keys(), [''] * len(TVMS_HEADERS.keys())))
                IP = ws[f'A{cell.row}'].value.split(",")
                result_list['F'] = VUL_NAME['M']
                result_list['G'] = FAILED
                result_list['K'] = "目前版本： " + ws[f'E{cell.row}'].value
                result_list['M'] = "更新至Microsoft Windows Server 2012或以上"
                result_list['O'] = CHECK_TYPE
                for ip in IP:
                    result_list['J'] = getRecSeq(count+1)
                    result_list['B'] = ip
                    result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                    output.append(",".join(list(result_list.values())))
                    count = count + 1
                result_list = None
    return output
def getMBSA(file_mbsa):
    MBSA = {}
    for f in file_mbsa.keys():
        ws = file_mbsa[f][0].active
        tmpdict = {}
        for row in ws.iter_rows(min_row=2):
            [IPs,latest,version,status] = [row[2].value,row[3].value,row[4].value,row[5].value]
            IPs = IPs.strip(";").split(";")
            for ip in IPs:
                tmpdict[ip] = {file_mbsa[f][1][0]:version,file_mbsa[f][1][1]:latest,file_mbsa[f][1][2]:status}
        MBSA[f] = tmpdict
    return MBSA
def getGCBSW(file_gcbsw):
    GCBSW = {}
    for filename in file_gcbsw:
        m = re.search('(?<=_)\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}.*(?=_)',filename)
        if m: # filename pattern "_[IPv4]_" found 
            ip = m.group(0)
            try:
                f = openpyxl.load_workbook(filename,read_only=False,data_only=True)
                ws = f.active
                office = []
                Antivirus = ""
                for row in ws.iter_rows(min_row=2):
                    if row[1].value:
                        if "Apex One Security Agent".upper() in row[1].value.upper():
                            Antivirus = row[1].value + "_" + row[2].value
                        elif "OfficeScan".upper() in row[1].value.upper() and len(Antivirus)==0:
                            Antivirus = row[1].value + "_" + row[2].value
                        elif row[0].value and ("Microsoft" in row[0].value) and ("Office" in row[1].value):
                            office.append(row[1].value + "_" + row[2].value)
                    
                    IPs = ip.split(",")
                    for p in IPs:
                        GCBSW[p] =  {MERGED_HEADER['D']:";".join(office),MERGED_HEADER['N']:Antivirus}
                f.close()
                
            except:
                print(f'Unable to open {filename}--', sys.exc_info()[0])
                raise
    for d in iter(GCBSW):
        if len(GCBSW[d]['Antivirus']) == 0:
            GCBSW[d]['Antivirus'] = "防毒軟體未安裝"
    return GCBSW
def getGCBOS(file_gcbos):
    GCBOS = {}
    ws = file_gcbos.active    
    for row in ws.iter_rows(min_row=2): # ignore the first row
        IPs = row[0].value.split(",")
        host_name = row[1].value
        os_version = row[4].value
        for ip in IPs:
            GCBOS[ip.strip()] = {MERGED_HEADER['A']:host_name,MERGED_HEADER['C']:os_version}
    return GCBOS
def getmerged_data(GCBOS,GCBSW,MBSA):
    merged_data={}
    headers = [x for x in MERGED_HEADER.values() if not x=="IP"]
    for d in GCBOS.items():
        ip = d[0]
        row = dict(zip(headers,[""]*len(headers)))
        row[MERGED_HEADER['A']] = d[1][MERGED_HEADER['A']]
        row[MERGED_HEADER['C']] = d[1][MERGED_HEADER['C']]
        
        if ip in GCBSW:
            row[MERGED_HEADER['D']] = GCBSW[ip][MERGED_HEADER['D']]
            row[MERGED_HEADER['N']] = GCBSW[ip][MERGED_HEADER['N']]
            
        if ip in MBSA['adobe']:
            row[MERGED_HEADER['E']] = MBSA['adobe'][ip][MERGED_HEADER['E']]
            row[MERGED_HEADER['F']] = MBSA['adobe'][ip][MERGED_HEADER['F']]
            row[MERGED_HEADER['G']] = MBSA['adobe'][ip][MERGED_HEADER['G']]
        if ip in MBSA['flashplayer']:
            row[MERGED_HEADER['H']] = MBSA['flashplayer'][ip][MERGED_HEADER['H']]
            row[MERGED_HEADER['I']] = MBSA['flashplayer'][ip][MERGED_HEADER['I']]
            row[MERGED_HEADER['J']] = MBSA['flashplayer'][ip][MERGED_HEADER['J']]
        if ip in MBSA['java']:
            row[MERGED_HEADER['K']] = MBSA['java'][ip][MERGED_HEADER['K']]
            row[MERGED_HEADER['L']] = MBSA['java'][ip][MERGED_HEADER['L']]
            row[MERGED_HEADER['M']] = MBSA['java'][ip][MERGED_HEADER['M']]
        merged_data[ip] = row
    return merged_data

def main():
    global AV_BASE_VER,RECORD_PREFIX
    parser = argparse.ArgumentParser()
    parser.add_argument("-r", "--read", help = "來源檔案路徑")
    parser.add_argument("-o", "--output", help = "輸出TVMS檔案路徑")
    parser.add_argument("-n", "--number",help = "評估工具原廠之弱點編號",default="")
    parser.add_argument("-a", "--version",help = "防毒軟體基準版本",default='14.0.9204')
    parser.add_argument("-m", "--merge",help = "產生彙整檔案",action="store_true",default=False)

    args = parser.parse_args()
    args = parser.parse_args()
    INPUT_DIR = os.path.abspath(args.read)
    OUTPUT_DIR      = args.output
    OUTPUT_FILE     = os.path.join(OUTPUT_DIR,"TVMS_Winserv_" + os.path.basename(INPUT_DIR)+".csv")
    AV_BASE_VER     = args.version
    RECORD_PREFIX   = args.number
    if args.merge:
        MERGED_FILE = os.path.join(OUTPUT_DIR,"MERGED_" + os.path.basename(INPUT_DIR)+".csv")
    else:
        MERGED_FILE = None
    
    init_input(INPUT_DIR) # read input files
    #開啟輸出彙整檔案
    if MERGED_FILE:
        try:
            merged_file = open(MERGED_FILE,encoding='utf-8-sig',mode='w')
            merged_file.write(",".join(list(MERGED_HEADER.values())) + '\n')
        except :
            print(f'Unable to open {MERGED_FILE}--', sys.exc_info()[0])
            raise
        MBSA = getMBSA({'adobe':        (R1,[MERGED_HEADER['E'],MERGED_HEADER['F'],MERGED_HEADER['G']]), \
                        'flashplayer':  (R2,[MERGED_HEADER['H'],MERGED_HEADER['I'],MERGED_HEADER['J']]), \
                        'java':         (R3,[MERGED_HEADER['K'],MERGED_HEADER['L'],MERGED_HEADER['M']])})
        GCBSW = getGCBSW(R5)
        GCBOS = getGCBOS(R6)
        merged_data = getmerged_data(GCBOS,GCBSW,MBSA)
        for ip,v in merged_data.items():
            line = list(v.values())
            line.insert(1, ip)
            merged_file.write(",".join(line) + '\n')
        merged_file.close()
    
    #開啟輸出檔案
    try:
        output_file = open(OUTPUT_FILE,encoding='utf-8-sig',mode='w')
    except :
        print(f'Unable to open {OUTPUT_FILE}--', sys.exc_info()[0])
        raise
    print(f'Output to : {OUTPUT_FILE}')
    output_file.write(",".join(list(TVMS_HEADERS.values()))+"\n")
    
    output = []
    #挑出【作業系統安全性更新編號】 & 【Office應用程式安全性更新編號】列
    output.extend(process_R0(R0))
    #挑出【Adobe Reader未更新】列
    output.extend(process_R1(R1))
    #挑出【Flash Player未更新】列
    output.extend(process_R2(R2))
    #挑出【Java未更新】列
    output.extend(process_R3(R3))
    #挑出【防毒軟體病毒碼未更新】& 【防毒軟體未安裝】列
    output.extend(process_R4(R4))
    # 挑出【防毒軟體未更新】& 【Office軟體版本過舊】列
    output.extend(process_R5(R4,R5))
    # 挑出【Windows作業系統版本過舊】列
    output.extend(process_R6(R6))
    output_file.write("\n".join(output))
    output_file.close()
    
if __name__ == '__main__':
    try:
        main()
    except:
        print(f'Failed--', sys.exc_info()[0])
        raise
