# -*- coding: utf-8 -*-
import openpyxl,sys,argparse
TVMS_HEADERS = {"A":"弱點發現時間",\
                "B":"弱點所在網路位址",\
                "C":"弱點所在網路埠號",\
                "D":"檔案名稱/URL",\
                "E":"弱點所在之網路協定",\
                "F":"弱點名稱",\
                "G":"弱點嚴重性或合規檢測結果",\
                "H":"弱點類別",\
                "I":"弱點CVE ID清單",\
                "J":"評估工具原廠之弱點編號",\
                "K":"弱點說明",\
                "L":"弱點意見",\
                "M":"弱點修補建議",\
                "N":"弱點證據描述",\
                "O":"類型(弱點或合規)"}
FAILED = "不符合"
PASSED = "符合"
TOOL_NAME = "神網資產管理系統健診說明"
CHECK_TYPE = "合規"
GCB_AUDIT = 'GCB合規檢測'

MAP2TVMS = {"B":"位址",\
            "G":"稽核結果",\
            "H":"類型",\
            "K":"組態名稱",\
            "M":"建議值",\
            "N":"本機設定值"}
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("-r", "--read", help = "來源Excel檔案路徑")
    parser.add_argument("-s", "--sheet", default = "主機分類總覽", help = "Excel工作表名稱")
    parser.add_argument("-o", "--output", help = "輸出TVMS檔案路徑")
    
    args = parser.parse_args()
    INPUT_FILE = args.read
    OUTPUT_FILE = args.output
    SHEET = args.sheet
    try:  #open input Excel file
        input_file = openpyxl.load_workbook(INPUT_FILE,read_only=False,data_only=True)
    except :
        print(f'Unable to open {INPUT_FILE}--', sys.exc_info()[0])
        raise
    if not input_file[SHEET]:
        sys.exit(f"檔案 :{INPUT_FILE} 內沒有工作表：{SHEET}")
    ws = input_file[SHEET]  # open desired worksheet in Excel
        
    try: # open the output file for writing
        output_file = open(OUTPUT_FILE,encoding='utf-8-sig',mode='w')
    except :
        print(f'Unable to open {OUTPUT_FILE}--', sys.exc_info()[0])
        raise
    output_file.write(",".join(list(TVMS_HEADERS.values()))+"\n") # the 1st line is the fields of TVMS  
    
    input_header = ws[1] #retrieve the 1st row of input row
    map2TVMS = {}
    for cell in input_header:
        try:
            map2TVMS[list(MAP2TVMS.keys())[list(MAP2TVMS.values()).index(cell.value)]] = cell.column_letter
        except ValueError: #cell.value is not a value in MAP2TVMS
            continue
       
    for cell in ws['K']: # lock to the column of result
        if cell.row == 1:
            continue # skip the first row
        else:
            if not (cell.value == PASSED): # only pick up the rows with failed result
                result_list = dict(zip(TVMS_HEADERS.keys(),['']*len(TVMS_HEADERS.keys())))
                for k,v in map2TVMS.items(): # map input file columns to TVMS column
                    try:
                        result_list[k] = ws[f'{v}{cell.row}'].value.replace("\n","").strip(" \n").replace(",","_")
                    except:
                        pass
                
                result_list['F'] = GCB_AUDIT
                result_list['G'] = FAILED
                result_list['O'] = CHECK_TYPE
                 
                result_list = dict(sorted(result_list.items(), key=lambda item: item[0]))
                output_file.write(",".join(list(result_list.values()))+"\n")
     
    input_file.close()
    output_file.close()
if __name__ == '__main__':
    try:
        main()
    except:
        print(f'Unable to run this program--', sys.exc_info()[0])
        raise
